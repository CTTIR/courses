"""Distribute VG Wort Zählmarken into the qmd files of a Quarto website.

See scripts/README_vgwort.md for usage. Hard rules: each marker is consumed
exactly once across all repos; xlsx is the source of truth and is updated
atomically alongside each qmd write; dry-run is the default.
"""

from __future__ import annotations

import argparse
import fnmatch
import os
import pathlib
import re
import shutil
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl pyyaml", file=sys.stderr)
    sys.exit(2)

try:
    import yaml
except ImportError:
    yaml = None  # frontmatter parsing falls back to regex


# ---------------------------------------------------------------------------
# Hard-coded exclusions — edit here, not via CLI
# ---------------------------------------------------------------------------

META_PAGE_NAMES = {
    "about.qmd",
    "impressum.qmd",
    "datenschutz.qmd",
    "datenschutzerklaerung.qmd",
    "kontakt.qmd",
    "contact.qmd",
    "imprint.qmd",
    "privacy.qmd",
    "licence.qmd",
    "license.qmd",
    "colophon.qmd",
}

META_PAGE_DIRS = {"about", "meta", "legal"}

DEFAULT_EXCLUDES = [
    "docs/**",
    "_freeze/**",
    ".quarto/**",
    "_extensions/**",
    "_inc/**",
    "_includes/**",
    "assets/**",
    ".git/**",
    "renv/**",
    "scripts/**",
]

INDEX_ONLY_THRESHOLD = 200       # chars; section index.qmd below this = INDEX-ONLY
HAUPTAUSSCHUETTUNG_MIN_CHARS = 1800  # informational only

VGWORT_HOST_RE = re.compile(r"vg(?:0\d|10)\.met\.vgwort\.de/na/([0-9a-fA-F]{32})")
VGWORT_DIV_RE = re.compile(r":::\s*\{\.vgwort-pixel\}")
VGWORT_COMMENT_RE = re.compile(r"VG Wort Zählmarke", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class XlsxRow:
    row_index: int  # 1-based incl. header
    public_id: str
    private_id: str
    html_code: str
    used: bool
    url: Optional[str]


@dataclass
class QmdFile:
    path: pathlib.Path
    rel_path: str
    classification: str  # META | INCLUDE | DRAFT | INDEX-ONLY | CONTENT
    classification_reason: str
    target_url: str
    has_existing_marker: bool
    existing_public_id: Optional[str]
    prose_chars: int


@dataclass
class RunStats:
    discovered: int = 0
    by_class: dict = field(default_factory=dict)
    already_marked_ok: int = 0
    reconciled: int = 0
    newly_allocated: int = 0
    below_threshold: int = 0
    anomalies: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Prose char counting (informational)
# ---------------------------------------------------------------------------

def count_prose_chars(text: str) -> int:
    text = re.sub(r"\A---\n.*?\n---\n", "", text, count=1, flags=re.DOTALL)
    text = re.sub(r"```.*?\n.*?```", "", text, flags=re.DOTALL)
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)
    text = re.sub(r"\$\$.*?\$\$", "", text, flags=re.DOTALL)
    text = re.sub(r"\$[^$\n]+\$", "", text)
    text = re.sub(r"\{\{<.*?>\}\}", "", text, flags=re.DOTALL)
    text = re.sub(r"`[^`\n]+`", "", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"^:::.*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s+", " ", text).strip()
    return len(text)


# ---------------------------------------------------------------------------
# Frontmatter parsing
# ---------------------------------------------------------------------------

def parse_frontmatter(text: str) -> dict:
    m = re.match(r"\A---\n(.*?)\n---\n", text, flags=re.DOTALL)
    if not m:
        return {}
    raw = m.group(1)
    if yaml is not None:
        try:
            data = yaml.safe_load(raw)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    out = {}
    for line in raw.splitlines():
        m2 = re.match(r"^([A-Za-z_][A-Za-z0-9_-]*)\s*:\s*(.+?)\s*$", line)
        if m2:
            k, v = m2.group(1), m2.group(2)
            if v.lower() in ("true", "false"):
                out[k] = v.lower() == "true"
            else:
                out[k] = v.strip("'\"")
    return out


# ---------------------------------------------------------------------------
# xlsx I/O with retry (Dropbox lock tolerance)
# ---------------------------------------------------------------------------

def with_retry(fn, *, attempts: int = 3, backoff: float = 2.0, what: str = "operation"):
    last = None
    for i in range(attempts):
        try:
            return fn()
        except (PermissionError, OSError) as exc:
            last = exc
            if i < attempts - 1:
                print(f"  [retry {i + 1}/{attempts}] {what}: {exc}", file=sys.stderr)
                time.sleep(backoff)
    raise RuntimeError(f"{what} failed after {attempts} attempts: {last}")


def load_xlsx(xlsx_path: pathlib.Path) -> tuple[openpyxl.Workbook, list[XlsxRow], dict]:
    wb = with_retry(lambda: openpyxl.load_workbook(xlsx_path), what="opening xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    expected = ["Öffentlicher Identifikationscode", "Privater Identifikationscode",
                "HTML_Code", "Used", "URL"]
    col = {name: header.index(name) for name in expected if name in header}
    missing = [n for n in expected if n not in col]
    if missing:
        raise RuntimeError(f"xlsx missing columns: {missing}")

    rows: list[XlsxRow] = []
    by_id: dict[str, XlsxRow] = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        pub = row[col["Öffentlicher Identifikationscode"]]
        if pub is None:
            continue
        pub = str(pub).strip().lower()
        priv = str(row[col["Privater Identifikationscode"]] or "").strip()
        html = str(row[col["HTML_Code"]] or "")
        used_raw = row[col["Used"]]
        if isinstance(used_raw, bool):
            used = used_raw
        else:
            used = str(used_raw).strip().lower() in ("true", "1", "yes", "x")
        url = row[col["URL"]]
        url = str(url) if url else None
        r = XlsxRow(row_index=i, public_id=pub, private_id=priv,
                    html_code=html, used=used, url=url)
        rows.append(r)
        by_id[pub] = r
    return wb, rows, by_id, col  # type: ignore[return-value]


def save_xlsx_atomic(wb: openpyxl.Workbook, xlsx_path: pathlib.Path) -> None:
    tmp = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")

    def _do():
        wb.save(tmp)
        # fsync on Windows: flush via opening in append-binary
        with open(tmp, "rb+") as f:
            os.fsync(f.fileno())
        os.replace(tmp, xlsx_path)

    with_retry(_do, what="saving xlsx")


def update_xlsx_row(wb: openpyxl.Workbook, col: dict, row: XlsxRow,
                    used: bool, url: Optional[str]) -> None:
    ws = wb.active
    ws.cell(row=row.row_index, column=col["Used"] + 1, value=used)
    ws.cell(row=row.row_index, column=col["URL"] + 1, value=url)
    row.used = used
    row.url = url


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------

def matches_any(rel: str, patterns: list[str]) -> bool:
    rel_norm = rel.replace("\\", "/")
    for p in patterns:
        if fnmatch.fnmatch(rel_norm, p):
            return True
    return False


def build_target_url(site_url: str, rel_path: str) -> str:
    rel = rel_path.replace("\\", "/").lstrip("./")
    if rel.endswith(".qmd"):
        rel = rel[:-4] + ".html"
    return site_url.rstrip("/") + "/" + rel


def detect_existing_marker(text: str) -> tuple[bool, Optional[str]]:
    m = VGWORT_HOST_RE.search(text)
    if m:
        return True, m.group(1).lower()
    if VGWORT_DIV_RE.search(text) or VGWORT_COMMENT_RE.search(text):
        return True, None
    return False, None


def classify(path: pathlib.Path, repo_root: pathlib.Path, text: str,
             prose_chars: int) -> tuple[str, str]:
    name = path.name
    rel = path.relative_to(repo_root).as_posix()
    parts = set(rel.split("/")[:-1])

    if name.startswith("_"):
        return "INCLUDE", "filename starts with _"

    fm = parse_frontmatter(text)
    if fm.get("vgwort") is False:
        return "META", "frontmatter vgwort: false"
    if fm.get("draft") is True:
        return "DRAFT", "frontmatter draft: true"

    if name in META_PAGE_NAMES:
        return "META", f"name in META_PAGE_NAMES ({name})"

    if name == "index.qmd" and path.parent == repo_root:
        return "META", "site-root index.qmd"

    if parts & META_PAGE_DIRS:
        return "META", f"directory in META_PAGE_DIRS"

    if name == "index.qmd" and prose_chars < INDEX_ONLY_THRESHOLD:
        return "INDEX-ONLY", f"section index.qmd with {prose_chars} chars (< {INDEX_ONLY_THRESHOLD})"

    return "CONTENT", ""


def discover(repo_root: pathlib.Path, site_url: str,
             excludes: list[str], onlys: list[str]) -> list[QmdFile]:
    files: list[QmdFile] = []
    for path in sorted(repo_root.rglob("*.qmd")):
        rel = path.relative_to(repo_root).as_posix()
        if matches_any(rel, excludes):
            continue
        if onlys and not matches_any(rel, onlys):
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except Exception as exc:
            print(f"WARN: cannot read {rel}: {exc}", file=sys.stderr)
            continue
        prose = count_prose_chars(text)
        cls, reason = classify(path, repo_root, text, prose)
        has_marker, existing_id = detect_existing_marker(text)
        files.append(QmdFile(
            path=path,
            rel_path=rel,
            classification=cls,
            classification_reason=reason,
            target_url=build_target_url(site_url, rel),
            has_existing_marker=has_marker,
            existing_public_id=existing_id,
            prose_chars=prose,
        ))
    return files


# ---------------------------------------------------------------------------
# Snippet rendering
# ---------------------------------------------------------------------------

def render_snippet(html_code: str, public_id: str) -> str:
    # xlsx HTML_Code stores literal \n escapes; convert to real newlines.
    img_html = html_code.replace("\\n", "\n").strip()
    if not img_html:
        # Fallback: build a canonical snippet from the public_id.
        img_html = (
            f'<img src="https://vg09.met.vgwort.de/na/{public_id}"\n'
            f'     width="1" height="1" alt=""\n'
            f'     loading="eager"\n'
            f'     fetchpriority="high"\n'
            f'     decoding="async"\n'
            f'     style="position:absolute;visibility:hidden;" />'
        )
    return (
        "\n"
        "::: {.vgwort-pixel}\n"
        "```{=html}\n"
        f"<!-- VG Wort Zählmarke — public ID: {public_id} -->\n"
        f"{img_html}\n"
        "```\n"
        ":::\n"
    )


def append_snippet_to_qmd(path: pathlib.Path, snippet: str) -> None:
    text = path.read_text(encoding="utf-8")
    if not text.endswith("\n"):
        text += "\n"
    text += snippet
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8", newline="\n")
    with open(tmp, "rb+") as f:
        os.fsync(f.fileno())
    os.replace(tmp, path)


# ---------------------------------------------------------------------------
# Main phases
# ---------------------------------------------------------------------------

def read_site_url(repo_root: pathlib.Path) -> str:
    qy = repo_root / "_quarto.yml"
    if not qy.exists():
        raise RuntimeError("_quarto.yml not found — not a Quarto project root?")
    text = qy.read_text(encoding="utf-8")
    m = re.search(r"site-url:\s*[\"']?([^\"'\n]+)[\"']?", text)
    if not m:
        raise RuntimeError("could not find website.site-url in _quarto.yml")
    return m.group(1).strip()


def assert_clean_git(repo_root: pathlib.Path) -> None:
    import subprocess
    r = subprocess.run(["git", "status", "--porcelain"], cwd=repo_root,
                       capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"git status failed: {r.stderr}")
    if r.stdout.strip():
        raise RuntimeError("working tree not clean — commit/stash before --apply:\n" + r.stdout)


def check_dropbox_conflict(xlsx_path: pathlib.Path) -> None:
    parent = xlsx_path.parent
    stem = xlsx_path.stem
    for sib in parent.glob(f"{stem}*CONFLICT*"):
        raise RuntimeError(f"Dropbox conflicted copy detected: {sib} — reconcile manually first")
    for sib in parent.glob(f"{stem}*conflicted copy*"):
        raise RuntimeError(f"Dropbox conflicted copy detected: {sib} — reconcile manually first")


def backup_xlsx(xlsx_path: pathlib.Path) -> pathlib.Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = xlsx_path.with_suffix(xlsx_path.suffix + f".bak.{ts}")
    shutil.copy2(xlsx_path, bak)
    return bak


def fmt_row(rel: str, cls: str, marked: bool, chars: int) -> str:
    return (f"  {rel:<70s}  class={cls:<10s}  "
            f"marked={'yes' if marked else 'no':<3s}  chars={chars:>5d}")


def run(args: argparse.Namespace) -> int:
    repo_root = pathlib.Path(args.root).resolve()
    xlsx_path = pathlib.Path(args.xlsx).resolve()

    print(f"Repo root: {repo_root}")
    print(f"xlsx:      {xlsx_path}")
    print(f"Mode:      {'APPLY' if args.apply else 'DRY RUN'}")
    print()

    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2
    check_dropbox_conflict(xlsx_path)

    site_url = read_site_url(repo_root)
    print(f"site-url:  {site_url}")

    if args.apply:
        assert_clean_git(repo_root)
        bak = backup_xlsx(xlsx_path)
        print(f"Backup:    {bak}")
    print()

    wb, rows, by_id, col = load_xlsx(xlsx_path)
    available = [r for r in rows if not r.used]
    print(f"Loaded xlsx: {len(rows)} rows, {len(available)} unused.")
    print()

    excludes = list(DEFAULT_EXCLUDES) + list(args.exclude or [])
    onlys = list(args.only or [])
    files = discover(repo_root, site_url, excludes, onlys)

    print(f"PHASE 1 — discovery ({len(files)} qmd files)")
    by_class: dict[str, list[QmdFile]] = {}
    for f in files:
        by_class.setdefault(f.classification, []).append(f)
        print(fmt_row(f.rel_path, f.classification, f.has_existing_marker, f.prose_chars))
    print()
    print("Counts by class:")
    for cls in ("CONTENT", "META", "INCLUDE", "DRAFT", "INDEX-ONLY"):
        print(f"  {cls:<12s} {len(by_class.get(cls, []))}")
    print()

    stats = RunStats(discovered=len(files),
                     by_class={k: len(v) for k, v in by_class.items()})

    # PHASE 2 — reconciliation
    print("PHASE 2 — reconciliation")
    xlsx_dirty = False
    for f in files:
        if not f.has_existing_marker:
            continue
        pid = f.existing_public_id
        if pid is None:
            stats.warnings.append(f"{f.rel_path}: marker detected but no public ID extractable")
            print(f"  WARN  {f.rel_path}: marker detected but no public ID extractable")
            continue
        row = by_id.get(pid)
        if row is None:
            stats.anomalies.append(f"{f.rel_path}: foreign marker {pid} (not in xlsx)")
            print(f"  ERROR {f.rel_path}: foreign marker {pid} not in xlsx — SKIPPING")
            continue
        if f.classification != "CONTENT":
            stats.warnings.append(
                f"{f.rel_path}: marker on non-content page (class={f.classification}); "
                f"reconciling xlsx but review manually")
            print(f"  WARN  {f.rel_path}: marker on {f.classification} page — review manually")
        if row.used and row.url == f.target_url:
            stats.already_marked_ok += 1
            print(f"  OK    {f.rel_path}  <-  {pid}")
        elif row.used and row.url != f.target_url:
            stats.warnings.append(
                f"{f.rel_path}: xlsx URL '{row.url}' != target '{f.target_url}' for {pid}")
            print(f"  WARN  {f.rel_path}: xlsx URL mismatch for {pid} "
                  f"(was {row.url!r}); not changing xlsx")
        else:
            update_xlsx_row(wb, col, row, used=True, url=f.target_url)
            xlsx_dirty = True
            stats.reconciled += 1
            print(f"  FIX   {f.rel_path}: xlsx Used=False -> True, URL set ({pid})")
    print()

    if stats.anomalies:
        print("ABORTING due to foreign markers (Phase 2 anomalies). "
              "Resolve manually before re-running.", file=sys.stderr)
        for a in stats.anomalies:
            print(f"  - {a}", file=sys.stderr)
        return 3

    # PHASE 3 — allocation
    print("PHASE 3 — allocation" + (" [DRY RUN]" if not args.apply else ""))
    available_iter = iter(r for r in rows if not r.used)
    to_allocate = [f for f in files
                   if f.classification == "CONTENT" and not f.has_existing_marker]

    if len(to_allocate) > len(available):
        print(f"ERROR: {len(to_allocate)} pages need markers but only "
              f"{len(available)} available — refusing to partially allocate.",
              file=sys.stderr)
        return 4

    for f in to_allocate:
        row = next(available_iter)
        snippet = render_snippet(row.html_code, row.public_id)
        if args.apply:
            append_snippet_to_qmd(f.path, snippet)
            update_xlsx_row(wb, col, row, used=True, url=f.target_url)
            save_xlsx_atomic(wb, xlsx_path)
            xlsx_dirty = False
            print(f"  OK    {f.rel_path}  <-  {row.public_id}  (chars={f.prose_chars})")
        else:
            print(f"  WOULD WRITE  {f.rel_path}  <-  {row.public_id}  (chars={f.prose_chars})")
        stats.newly_allocated += 1
        if f.prose_chars < HAUPTAUSSCHUETTUNG_MIN_CHARS:
            stats.below_threshold += 1
    print()

    # If we only did Phase-2 reconciliation, save now.
    if args.apply and xlsx_dirty:
        save_xlsx_atomic(wb, xlsx_path)

    # PHASE 4 — exclusion / informational reports
    print("PHASE 4 — exclusions (informational)")
    for cls in ("META", "INCLUDE", "DRAFT", "INDEX-ONLY"):
        bucket = by_class.get(cls, [])
        if not bucket:
            continue
        print(f"  [{cls}]")
        for f in bucket:
            print(f"    {f.rel_path}  ({f.classification_reason})")
    print()
    short_marked = [f for f in to_allocate if f.prose_chars < HAUPTAUSSCHUETTUNG_MIN_CHARS]
    if short_marked:
        print(f"  Below Hauptausschüttung threshold ({HAUPTAUSSCHUETTUNG_MIN_CHARS} chars) — "
              f"marker placed, but consider expanding for full payout eligibility:")
        for f in short_marked:
            print(f"    {f.rel_path}  chars={f.prose_chars}")
        print()

    # PHASE 5 — summary
    remaining = sum(1 for r in rows if not r.used)
    print("SUMMARY")
    print(f"  Discovered:                  {stats.discovered}")
    print(f"  Excluded (META):             {stats.by_class.get('META', 0)}")
    print(f"  Excluded (INCLUDE):          {stats.by_class.get('INCLUDE', 0)}")
    print(f"  Excluded (DRAFT):            {stats.by_class.get('DRAFT', 0)}")
    print(f"  Excluded (INDEX-ONLY):       {stats.by_class.get('INDEX-ONLY', 0)}")
    print(f"  CONTENT pages:               {stats.by_class.get('CONTENT', 0)}")
    print(f"    Already marked (OK):       {stats.already_marked_ok}")
    print(f"    Reconciled (xlsx fixed):   {stats.reconciled}")
    print(f"    Newly allocated:           {stats.newly_allocated}")
    print(f"      of which < {HAUPTAUSSCHUETTUNG_MIN_CHARS} chars:    {stats.below_threshold}")
    print(f"  Warnings:                    {len(stats.warnings)}")
    print(f"  Anomalies:                   {len(stats.anomalies)}")
    print(f"  Available remaining:         {remaining}")
    if stats.warnings:
        print()
        print("  Warnings:")
        for w in stats.warnings:
            print(f"    - {w}")
    print()
    if not args.apply:
        print("DRY RUN — re-run with --apply to write changes")
    else:
        print("DONE — review with `git diff` and commit")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, help="Path to zaehlmarken_combined.xlsx")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--apply", action="store_true", help="Write changes (default: dry-run)")
    g.add_argument("--dry-run", action="store_true", help="Show plan without writing (default)")
    ap.add_argument("--root", default=".", help="Repo root (default: cwd)")
    ap.add_argument("--exclude", action="append", default=[],
                    help="Additional path globs to exclude (relative, posix)")
    ap.add_argument("--only", action="append", default=[],
                    help="Restrict to matching path globs (relative, posix)")
    args = ap.parse_args()
    try:
        return run(args)
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
